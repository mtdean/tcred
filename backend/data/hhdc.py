"""
backend/data/hhdc.py — NY Fed Household Debt & Credit Report ingest.

Pulls the quarterly "flow into delinquency" transition rates by loan type from
the NY Fed's published data workbook. These flows (% of balances transitioning
*into* 30+ / 90+ days delinquent) lead the stock delinquency/charge-off series
we already track from FRED — flows turn before stocks.

Source: the NY Fed Consumer Credit Panel / Equifax. The workbook is published
~6 weeks after quarter-end at a predictable URL; we probe recent quarters and
take the newest that exists. Stored in the `metrics` table under
`delinquency_flow`, so they surface through the existing FRED endpoints.

Data layout (workbook tabs are page-numbered, mapped via the table of contents):
  Page 13 Data — "Flow into Early Delinquency (30+) by Loan Type"
  Page 14 Data — "Flow into Serious Delinquency (90+) by Loan Type"
Each: header row with AUTO/CC/MORTGAGE/HELOC/STUDENT LOAN/OTHER/Total, then
rows keyed by quarter label "YY:Qn" (e.g. "26:Q1" = 2026 Q1).
"""

import io
import logging
from datetime import date, datetime, timezone

import requests

from cache.db import upsert_metric

logger = logging.getLogger(__name__)

_BASE_URL = (
    "https://www.newyorkfed.org/medialibrary/interactives/"
    "householdcredit/data/xls/hhd_c_report_{year}q{q}.xlsx"
)
# The NY Fed CDN 403s bare clients; a browser-ish UA with contact is accepted.
_HEADERS = {"User-Agent": "Mozilla/5.0 SituationMonitor/0.1 mtuckerdean@gmail.com"}

# (workbook tab, delinquency bucket) for the two flow pages.
_FLOW_PAGES = [("Page 13 Data", 30), ("Page 14 Data", 90)]

# Workbook column header → stable series-id suffix.
_LOAN_KEYS = {
    "AUTO": "AUTO",
    "CC": "CC",
    "MORTGAGE": "MORTGAGE",
    "HELOC": "HELOC",
    "STUDENT LOAN": "STUDENT",
    "OTHER": "OTHER",
    "TOTAL": "ALL",
    "ALL": "ALL",
}
_LOAN_LABEL = {
    "AUTO": "Auto",
    "CC": "Credit Card",
    "MORTGAGE": "Mortgage",
    "HELOC": "HELOC",
    "STUDENT": "Student Loan",
    "OTHER": "Other",
    "ALL": "All Loans",
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _recent_quarters(n: int = 6) -> list[tuple[int, int]]:
    """The last `n` calendar quarters, newest first, as (year, quarter)."""
    today = date.today()
    y, q = today.year, (today.month - 1) // 3 + 1
    out: list[tuple[int, int]] = []
    for _ in range(n):
        out.append((y, q))
        q -= 1
        if q == 0:
            q, y = 4, y - 1
    return out


def _quarter_label_to_date(label: str) -> str | None:
    """'26:Q1' -> '2026-01-01'. Quarter start, matching FRED's convention."""
    label = label.strip().upper()
    if ":Q" not in label:
        return None
    yy, q = label.split(":Q")
    try:
        yy_i, q_i = int(yy), int(q)
    except ValueError:
        return None
    if not 1 <= q_i <= 4:
        return None
    year = 2000 + yy_i if yy_i < 50 else 1900 + yy_i
    month = (q_i - 1) * 3 + 1
    return f"{year:04d}-{month:02d}-01"


def _fetch_latest_workbook() -> tuple[bytes, str] | None:
    """Download the newest available HHDC workbook. Returns (bytes, 'YYYYQn')."""
    for year, q in _recent_quarters():
        url = _BASE_URL.format(year=year, q=q)
        try:
            resp = requests.get(url, headers=_HEADERS, timeout=60)
        except Exception as e:
            logger.warning("HHDC probe %dQ%d error: %s", year, q, e)
            continue
        # Real workbooks are ~1MB; a 200 with a tiny body is an error page.
        if resp.status_code == 200 and len(resp.content) > 100_000:
            logger.info("HHDC: using %dQ%d (%d bytes)", year, q, len(resp.content))
            return resp.content, f"{year}Q{q}"
    logger.error("HHDC: no workbook found in the last %d quarters", len(_recent_quarters()))
    return None


def _parse_flow_sheet(ws, bucket: int, now: str) -> int:
    """Parse one flow page into per-loan-type metrics. Returns rows written."""
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row: the one whose cells include the loan-type names.
    header_idx = None
    for i, row in enumerate(rows):
        labels = {str(c).strip().upper() for c in row if c is not None}
        if "AUTO" in labels and "MORTGAGE" in labels:
            header_idx = i
            break
    if header_idx is None:
        logger.warning("HHDC: no header row in %s", ws.title)
        return 0

    header = rows[header_idx]
    # Map column index -> loan key (skip the date column and blanks).
    col_keys: dict[int, str] = {}
    for ci, cell in enumerate(header):
        if cell is None:
            continue
        key = _LOAN_KEYS.get(str(cell).strip().upper())
        if key:
            col_keys[ci] = key

    count = 0
    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        d = _quarter_label_to_date(str(row[0]))
        if not d:
            continue
        for ci, key in col_keys.items():
            if ci >= len(row):
                continue
            val = row[ci]
            if val is None or isinstance(val, str):
                continue
            upsert_metric(
                {
                    "series_id": f"HHDC_FLOW{bucket}_{key}",
                    "label": f"Flow into {bucket}+ Delinquency — {_LOAN_LABEL[key]} (NY Fed)",
                    "category": "delinquency_flow",
                    "date": d,
                    "value": round(float(val), 4),
                    "fetched_at": now,
                }
            )
            count += 1
    return count


def fetch_hhdc_transitions() -> int:
    """Fetch and store NY Fed flow-into-delinquency rates. Returns rows written."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return 0

    fetched = _fetch_latest_workbook()
    if not fetched:
        return 0
    content, _quarter = fetched

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.error("HHDC: workbook parse error: %s", e)
        return 0

    now = _now()
    total = 0
    for tab, bucket in _FLOW_PAGES:
        if tab not in wb.sheetnames:
            logger.warning("HHDC: tab %r missing", tab)
            continue
        total += _parse_flow_sheet(wb[tab], bucket, now)

    logger.info("HHDC: stored %d transition-rate points", total)
    return total
