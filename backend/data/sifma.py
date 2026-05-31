"""
backend/data/sifma.py — SIFMA ABS issuance pipeline.

SIFMA's downloadable Excel is behind an HSForms gate, so auto-download isn't
practical. Workflow instead:

  1. User downloads the latest "US ABS Issuance" xlsx from sifma.org once a
     month and drops it into `backend/cache/sifma_drops/`.
  2. The scheduled `_job_sifma` (every 6h) scans the drop folder, parses any
     new file, ingests rows into the `metrics` table under series ids
     SIFMA_AUTO / SIFMA_CC / SIFMA_STUDENT / SIFMA_EQUIP / SIFMA_HE /
     SIFMA_MH / SIFMA_OTHER / SIFMA_TOTAL, and moves the parsed file to
     `backend/cache/sifma_drops/parsed/`.
  3. The /api/abs/issuance endpoint serves whatever has landed.

The parser is intentionally tolerant of SIFMA's exact sheet layout — it
detects the header row by scanning for asset-class keywords rather than
assuming a fixed cell position, so minor reformatting on SIFMA's side
shouldn't break us. If the parser can't find a header row it logs and skips.
Values are stored as USD billions to match SIFMA's published units.
"""

from __future__ import annotations

import logging
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional

from cache import db
from config import settings

logger = logging.getLogger(__name__)


DROP_DIR = settings.DB_PATH.parent / "sifma_drops"
ARCHIVE_DIR = DROP_DIR / "parsed"


# Header-keyword → series_id. The parser matches header cells case-insensitively
# against any keyword in the list; first hit wins per column. Designed for
# SIFMA's documented asset-class categories.
_ASSET_CLASS_KEYWORDS: list[tuple[str, list[str]]] = [
    ("SIFMA_AUTO",    ["auto"]),
    ("SIFMA_CC",      ["credit card", "credit-card", "credit cards"]),
    ("SIFMA_STUDENT", ["student"]),
    ("SIFMA_EQUIP",   ["equipment"]),
    ("SIFMA_HE",      ["home equity", "home-equity", "heloc"]),
    ("SIFMA_MH",      ["manufactured housing", "manufactured-housing"]),
    ("SIFMA_OTHER",   ["other"]),
    ("SIFMA_TOTAL",   ["total"]),
]

_LABELS = {
    "SIFMA_AUTO":    "SIFMA US ABS Issuance — Auto ($B)",
    "SIFMA_CC":      "SIFMA US ABS Issuance — Credit Cards ($B)",
    "SIFMA_STUDENT": "SIFMA US ABS Issuance — Student Loans ($B)",
    "SIFMA_EQUIP":   "SIFMA US ABS Issuance — Equipment ($B)",
    "SIFMA_HE":      "SIFMA US ABS Issuance — Home Equity ($B)",
    "SIFMA_MH":      "SIFMA US ABS Issuance — Manufactured Housing ($B)",
    "SIFMA_OTHER":   "SIFMA US ABS Issuance — Other ($B)",
    "SIFMA_TOTAL":   "SIFMA US ABS Issuance — Total ($B)",
}

CATEGORY = "abs_issuance"


def _now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


# ── Header detection ────────────────────────────────────────────────────────

def _classify_header(cells: list[str]) -> dict[int, str]:
    """Map a header row's column index → series_id, by keyword match.

    A cell matches a series_id if any of its keywords appears as a substring
    in the normalized cell text. First-match-wins per series_id (so if both
    "Credit Cards" and "Credit Card" appear, only the first column gets
    assigned to SIFMA_CC).
    """
    out: dict[int, str] = {}
    used: set[str] = set()
    for i, raw in enumerate(cells):
        n = _norm(raw)
        if not n:
            continue
        for sid, keywords in _ASSET_CLASS_KEYWORDS:
            if sid in used:
                continue
            if any(kw in n for kw in keywords):
                # Skip pure "other" cells that are obviously a subtotal label
                # rather than the Other asset class (heuristic safety).
                if sid == "SIFMA_OTHER" and n == "other" and i == 0:
                    continue
                out[i] = sid
                used.add(sid)
                break
    return out


_DATE_HEADER_KEYWORDS = ("date", "month", "year", "period")


def _is_date_cell(text: str) -> bool:
    if not text:
        return False
    n = _norm(text)
    return any(kw in n for kw in _DATE_HEADER_KEYWORDS)


def _find_header_row(rows: list[list[str]]) -> Optional[int]:
    """Pick the first row that maps ≥ 3 asset-class columns. Cheap heuristic
    that survives SIFMA prefacing the sheet with title rows / footnotes."""
    for i, row in enumerate(rows[:60]):  # cap scan; data starts within ~50 rows
        if len(_classify_header(row)) >= 3:
            return i
    return None


# ── Date parsing ────────────────────────────────────────────────────────────

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10,
    "november": 11, "december": 12,
}


def _parse_period(raw) -> Optional[str]:
    """Coerce a cell value into a quarter-start ISO date 'YYYY-MM-01'.

    Accepts: datetime, '2026-04', '2026/04', '2026-04-15', 'Apr 2026',
    '2026 Apr', 'April 2026', integer-year (treated as Jan), 'YYYY-Mn' etc.
    Returns None for anything we can't interpret confidently.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return f"{raw.year:04d}-{raw.month:02d}-01"
    if isinstance(raw, (int, float)):
        # An integer year cell — common in SIFMA's "Annual" rows. Treat as
        # January of that year for storage.
        try:
            y = int(raw)
            if 1980 <= y <= 2100:
                return f"{y:04d}-01-01"
        except (TypeError, ValueError):
            pass
        return None

    s = _norm(raw)
    if not s:
        return None

    # YYYY-MM, YYYY/MM, YYYY-MM-DD, YYYY.MM
    m = re.match(r"^(\d{4})[-/.](\d{1,2})(?:[-/.]\d{1,2})?$", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}-{mo:02d}-01"

    # Month name + year (e.g. "apr 2026", "april 2026", "2026 apr").
    for fmt in (
        re.compile(r"^([a-z]{3,10})\s+(\d{4})$"),
        re.compile(r"^(\d{4})\s+([a-z]{3,10})$"),
    ):
        m = fmt.match(s)
        if m:
            a, b = m.group(1), m.group(2)
            month_part = a if a.isalpha() else b
            year_part = b if a.isalpha() else a
            mo = _MONTHS.get(month_part)
            if mo:
                return f"{int(year_part):04d}-{mo:02d}-01"

    # Bare 4-digit year.
    m = re.match(r"^(\d{4})$", s)
    if m:
        y = int(m.group(1))
        if 1980 <= y <= 2100:
            return f"{y:04d}-01-01"
    return None


# ── xlsx parsing ────────────────────────────────────────────────────────────

def _sheet_to_rows(ws) -> list[list[str]]:
    return [
        [(cell.value if cell.value is not None else "") for cell in row]
        for row in ws.iter_rows()
    ]


def _coerce_value(raw) -> Optional[float]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "")
    if s in ("", "-", "—", "n/a", "N/A", "NA"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_sifma_xlsx(path: Path | str) -> list[dict]:
    """Parse one SIFMA xlsx, returning per (date, series_id, value) records.

    Iterates every sheet, picks the first one whose header row classifies at
    least three asset-class columns. The date column is whichever column
    yields parseable `_parse_period(...)` values for most data rows.
    """
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows = _sheet_to_rows(ws)
            hdr_idx = _find_header_row(rows)
            if hdr_idx is None:
                continue
            colmap = _classify_header(rows[hdr_idx])
            data_rows = rows[hdr_idx + 1:]

            # Pick the date column: the column whose values parse as a period
            # for the largest number of rows. Tie-break: prefer a column whose
            # header contains 'date'/'month'/'year'/'period'.
            n_cols = max(len(r) for r in data_rows) if data_rows else 0
            best_col = None
            best_hits = 0
            for c in range(n_cols):
                if c in colmap:
                    continue  # an asset-class column can't also be the date col
                hits = sum(
                    1 for r in data_rows
                    if c < len(r) and _parse_period(r[c]) is not None
                )
                # Boost columns whose header says "date" / "month" / "year".
                header_cell = rows[hdr_idx][c] if c < len(rows[hdr_idx]) else ""
                if _is_date_cell(str(header_cell)):
                    hits += 5
                if hits > best_hits:
                    best_hits = hits
                    best_col = c
            if best_col is None or best_hits == 0:
                continue

            now = _now_utc_iso()
            out: list[dict] = []
            for row in data_rows:
                if best_col >= len(row):
                    continue
                period = _parse_period(row[best_col])
                if not period:
                    continue
                for col_idx, sid in colmap.items():
                    if col_idx >= len(row):
                        continue
                    val = _coerce_value(row[col_idx])
                    if val is None:
                        continue
                    out.append({
                        "series_id": sid,
                        "label": _LABELS[sid],
                        "category": CATEGORY,
                        "date": period,
                        "value": val,
                        "fetched_at": now,
                    })
            if out:
                logger.info(
                    "SIFMA parse %s: %d records (sheet=%r, asset_classes=%s)",
                    path.name, len(out), ws.title, sorted(set(colmap.values())),
                )
                return out
    finally:
        wb.close()
    logger.warning("SIFMA parse %s: no recognizable issuance sheet", path.name)
    return []


# ── Drop folder ingest ──────────────────────────────────────────────────────

def _ensure_dirs() -> None:
    DROP_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)


def _new_xlsx_drops() -> list[Path]:
    if not DROP_DIR.is_dir():
        return []
    out = []
    for p in DROP_DIR.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        if p.name.startswith("."):  # hide macOS .DS_Store + tmp files
            continue
        out.append(p)
    return sorted(out)


def _archive(path: Path) -> Path:
    """Move a parsed file into the archive dir, prefixing with a UTC stamp
    so re-dropping the same filename doesn't overwrite history."""
    _ensure_dirs()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    dest = ARCHIVE_DIR / f"{stamp}__{path.name}"
    shutil.move(str(path), str(dest))
    return dest


def ingest_sifma_drops() -> dict:
    """Scan DROP_DIR, parse + ingest each new xlsx, archive when done.

    Idempotent: if no new files have been dropped, returns counts of zero.
    Failed parses (zero records) still archive the file so we don't keep
    retrying the same bad input every cycle — the file lands in
    `parsed/__failed/` instead so the user can inspect.
    """
    _ensure_dirs()
    drops = _new_xlsx_drops()
    if not drops:
        return {"files": 0, "records": 0, "rows_written": 0}

    failed_dir = ARCHIVE_DIR / "__failed"
    failed_dir.mkdir(parents=True, exist_ok=True)

    files = 0
    records = 0
    rows_written = 0
    for path in drops:
        files += 1
        try:
            recs = parse_sifma_xlsx(path)
        except Exception as e:  # noqa: BLE001
            logger.error("SIFMA parse error on %s: %s", path.name, e)
            recs = []

        if not recs:
            stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            shutil.move(str(path), str(failed_dir / f"{stamp}__{path.name}"))
            continue

        for r in recs:
            db.upsert_metric(r)
            rows_written += 1
        records += len(recs)
        _archive(path)

    db.set_meta("last_sifma_ingest", _now_utc_iso())
    logger.info(
        "SIFMA ingest: %d files, %d records, %d rows written",
        files, records, rows_written,
    )
    return {
        "files": files,
        "records": records,
        "rows_written": rows_written,
        "drop_dir": str(DROP_DIR),
    }


# ── Optional FRED supplement ────────────────────────────────────────────────

# Coarser, automatic complement to SIFMA. FRED publishes an aggregate ABS
# series (no per-class breakdown). If the configured id isn't a real FRED
# series the fetch logs a warning and the panel just hides that line.
_FRED_SUPPLEMENT_SERIES = "ABSI"  # configurable in data_sources.yaml if needed
_FRED_SUPPLEMENT_LABEL = "FRED ABS Issuance (aggregate)"
_FRED_SUPPLEMENT_ID = "FRED_ABS_ISSUANCE"


def fetch_fred_abs_issuance() -> int:
    """Pull a FRED aggregate ABS issuance series, store under FRED_ABS_ISSUANCE."""
    if not settings.FRED_API_KEY:
        return 0
    try:
        from fredapi import Fred
        fred = Fred(api_key=settings.FRED_API_KEY)
        series = fred.get_series(_FRED_SUPPLEMENT_SERIES, observation_start="2005-01-01").dropna()
    except Exception as e:  # noqa: BLE001 — FRED 404 / SDK / network all collapse here
        logger.info("FRED ABS issuance supplement skipped (%s)", e)
        return 0

    now = _now_utc_iso()
    n = 0
    for d, v in series.items():
        db.upsert_metric({
            "series_id": _FRED_SUPPLEMENT_ID,
            "label": _FRED_SUPPLEMENT_LABEL,
            "category": CATEGORY,
            "date": str(d.date()),
            "value": float(v),
            "fetched_at": now,
        })
        n += 1
    return n


# ── Query helpers ───────────────────────────────────────────────────────────

ALL_SIFMA_IDS = [sid for sid, _ in _ASSET_CLASS_KEYWORDS]


def get_issuance_series() -> dict:
    """Return per-asset-class monthly issuance plus the FRED supplement."""
    with db.get_conn() as conn:
        rows = conn.execute(
            "SELECT series_id, date, value FROM metrics "
            "WHERE category = ? AND value IS NOT NULL ORDER BY date ASC",
            (CATEGORY,),
        ).fetchall()
    by_series: dict[str, list[dict]] = {}
    for r in rows:
        by_series.setdefault(r["series_id"], []).append(
            {"date": r["date"], "value": float(r["value"])}
        )
    return {
        "sifma": {
            sid: {
                "label": _LABELS.get(sid, sid),
                "points": by_series.get(sid, []),
            }
            for sid in ALL_SIFMA_IDS
            if by_series.get(sid)
        },
        "fred_supplement": {
            "series_id": _FRED_SUPPLEMENT_ID,
            "label": _FRED_SUPPLEMENT_LABEL,
            "points": by_series.get(_FRED_SUPPLEMENT_ID, []),
        },
        "last_ingest": db.get_meta("last_sifma_ingest"),
        "drop_dir": str(DROP_DIR),
    }
