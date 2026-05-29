"""
Cover data/hhdc.py — NY Fed Household Debt & Credit transitions ingest.

Strategy:
  * Pure helpers (_quarter_label_to_date, _recent_quarters) get table tests.
  * End-to-end fetch path uses `mocked_responses` to serve an in-memory
    openpyxl workbook for the newest probed quarter, and asserts the right
    rows landed in metrics.
  * 404 / empty / network failures return 0 without raising.
"""

from __future__ import annotations

import io
from datetime import date

import pytest

from cache import db
from data import hhdc


# ─── _quarter_label_to_date ───────────────────────────────────────────────────
class TestQuarterLabelToDate:
    @pytest.mark.parametrize("label, expected", [
        ("26:Q1", "2026-01-01"),
        ("26:Q2", "2026-04-01"),
        ("26:Q3", "2026-07-01"),
        ("26:Q4", "2026-10-01"),
        ("00:Q1", "2000-01-01"),
        ("99:Q4", "1999-10-01"),  # < 50 → 21st century, >= 50 → 20th century
        ("49:Q1", "2049-01-01"),
        ("50:Q1", "1950-01-01"),
    ])
    def test_valid_labels(self, label, expected):
        assert hhdc._quarter_label_to_date(label) == expected

    @pytest.mark.parametrize("label", [
        "26Q1",           # missing colon
        "26:Q5",          # bad quarter
        "26:Q0",          # bad quarter
        "AB:Q1",          # non-numeric year
        "26:QX",          # non-numeric quarter
        "random string",
        "",
    ])
    def test_invalid_labels(self, label):
        assert hhdc._quarter_label_to_date(label) is None

    def test_strips_whitespace_and_uppercases(self):
        # The function does `strip()` + `upper()` so lower-case lables work too.
        assert hhdc._quarter_label_to_date("  26:q3  ") == "2026-07-01"


# ─── _recent_quarters ─────────────────────────────────────────────────────────
class TestRecentQuarters:
    def test_returns_n_quarters_newest_first(self):
        out = hhdc._recent_quarters(4)
        assert len(out) == 4
        # Each entry is a (year, quarter) tuple.
        for y, q in out:
            assert 1 <= q <= 4
            assert y >= 1990

    def test_walks_back_one_quarter_at_a_time(self):
        out = hhdc._recent_quarters(8)
        # Convert to an ordinal and verify it's strictly decreasing by 1.
        ords = [y * 4 + (q - 1) for y, q in out]
        assert all(ords[i] - ords[i + 1] == 1 for i in range(len(ords) - 1))

    def test_default_n_is_six(self):
        assert len(hhdc._recent_quarters()) == 6

    def test_first_entry_is_current_quarter(self):
        out = hhdc._recent_quarters(1)
        today = date.today()
        expected_q = (today.month - 1) // 3 + 1
        assert out[0] == (today.year, expected_q)


# ─── _now string ──────────────────────────────────────────────────────────────
class TestNow:
    def test_iso_format(self):
        s = hhdc._now()
        # Best-effort: parseable by fromisoformat.
        from datetime import datetime
        datetime.fromisoformat(s)


# ─── _fetch_latest_workbook ───────────────────────────────────────────────────
def _make_minimal_workbook() -> bytes:
    """Build a real .xlsx with the two flow pages and one quarter row each."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Page 13 → 30+ flow.
    ws30 = wb.create_sheet("Page 13 Data")
    # Junk header row to exercise the "find header row" logic.
    ws30.append(["Flow into Early Delinquency (30+) by Loan Type"])
    ws30.append(["Date", "AUTO", "CC", "MORTGAGE", "HELOC",
                 "STUDENT LOAN", "OTHER", "TOTAL"])
    ws30.append(["26:Q1", 1.5, 8.0, 1.0, 0.5, 9.0, 2.0, 3.0])
    ws30.append(["26:Q2", 1.6, 8.5, 1.1, 0.55, 9.5, 2.1, 3.2])

    ws90 = wb.create_sheet("Page 14 Data")
    ws90.append(["Flow into Serious Delinquency (90+) by Loan Type"])
    ws90.append(["Date", "AUTO", "CC", "MORTGAGE", "HELOC",
                 "STUDENT LOAN", "OTHER", "TOTAL"])
    ws90.append(["26:Q1", 0.5, 4.0, 0.4, 0.2, 6.0, 1.0, 1.5])

    # Pad to >100kB so the size sniff doesn't reject the workbook.
    pad = wb.create_sheet("padding")
    for _ in range(20000):
        pad.append(["x"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestFetchLatestWorkbook:
    def test_returns_first_quarter_with_real_payload(
        self, fresh_db, mocked_responses
    ):
        wb_bytes = _make_minimal_workbook()
        quarters = hhdc._recent_quarters()
        # First quarter (newest) returns a 200 with sufficient body.
        first_y, first_q = quarters[0]
        url = hhdc._BASE_URL.format(year=first_y, q=first_q)
        mocked_responses.get(url, body=wb_bytes, status=200)
        result = hhdc._fetch_latest_workbook()
        assert result is not None
        content, q_label = result
        assert q_label == f"{first_y}Q{first_q}"
        assert len(content) > 100_000

    def test_falls_through_to_next_quarter_when_404(
        self, fresh_db, mocked_responses
    ):
        wb_bytes = _make_minimal_workbook()
        quarters = hhdc._recent_quarters()
        # Newest returns 404, second-newest returns the real workbook.
        y0, q0 = quarters[0]
        y1, q1 = quarters[1]
        mocked_responses.get(
            hhdc._BASE_URL.format(year=y0, q=q0), status=404,
        )
        mocked_responses.get(
            hhdc._BASE_URL.format(year=y1, q=q1), body=wb_bytes, status=200,
        )
        result = hhdc._fetch_latest_workbook()
        assert result is not None
        _, q_label = result
        assert q_label == f"{y1}Q{q1}"

    def test_returns_none_when_all_quarters_fail(
        self, fresh_db, mocked_responses
    ):
        for y, q in hhdc._recent_quarters():
            mocked_responses.get(
                hhdc._BASE_URL.format(year=y, q=q), status=404,
            )
        assert hhdc._fetch_latest_workbook() is None

    def test_tiny_response_treated_as_error_page(
        self, fresh_db, mocked_responses
    ):
        # All quarters return a 200 but with a body smaller than the
        # 100,000 byte threshold → treated as error pages → None.
        for y, q in hhdc._recent_quarters():
            mocked_responses.get(
                hhdc._BASE_URL.format(year=y, q=q),
                body=b"<html>error</html>", status=200,
            )
        assert hhdc._fetch_latest_workbook() is None


# ─── fetch_hhdc_transitions: end-to-end ───────────────────────────────────────
class TestFetchHhdcTransitions:
    def test_persists_rows_for_both_flow_buckets(
        self, fresh_db, mocked_responses
    ):
        wb_bytes = _make_minimal_workbook()
        y0, q0 = hhdc._recent_quarters()[0]
        mocked_responses.get(
            hhdc._BASE_URL.format(year=y0, q=q0), body=wb_bytes, status=200,
        )
        # All older probes also 404 — they should never be hit since the
        # first probe succeeded, but stub them to be safe.
        for y, q in hhdc._recent_quarters()[1:]:
            mocked_responses.get(
                hhdc._BASE_URL.format(year=y, q=q), status=404,
            )

        n = hhdc.fetch_hhdc_transitions()
        # Page13 has 2 rows × 7 loan keys (AUTO, CC, MORTGAGE, HELOC, STUDENT,
        # OTHER, TOTAL→ALL) = 14; Page14 has 1 × 7 = 7. Total = 21.
        assert n == 21

        # Sample a couple of writes by series_id.
        with db.get_conn() as conn:
            auto30 = conn.execute(
                "SELECT date, value FROM metrics "
                "WHERE series_id='HHDC_FLOW30_AUTO' ORDER BY date"
            ).fetchall()
        assert [(r["date"], r["value"]) for r in auto30] == [
            ("2026-01-01", 1.5),
            ("2026-04-01", 1.6),
        ]

        # "TOTAL" column maps to "ALL" suffix per _LOAN_KEYS.
        with db.get_conn() as conn:
            all_rows = conn.execute(
                "SELECT date, value FROM metrics "
                "WHERE series_id='HHDC_FLOW90_ALL'"
            ).fetchall()
        assert len(all_rows) == 1
        assert all_rows[0]["value"] == 1.5

    def test_no_workbook_found_returns_zero(self, fresh_db, mocked_responses):
        for y, q in hhdc._recent_quarters():
            mocked_responses.get(
                hhdc._BASE_URL.format(year=y, q=q), status=404,
            )
        assert hhdc.fetch_hhdc_transitions() == 0

    def test_corrupt_workbook_returns_zero(self, fresh_db, mocked_responses):
        # Send back a > 100kB blob that openpyxl can't parse → caught,
        # function returns 0 without raising.
        y0, q0 = hhdc._recent_quarters()[0]
        garbage = b"X" * 200_000
        mocked_responses.get(
            hhdc._BASE_URL.format(year=y0, q=q0), body=garbage, status=200,
        )
        assert hhdc.fetch_hhdc_transitions() == 0


# ─── _parse_flow_sheet directly ───────────────────────────────────────────────
class TestParseFlowSheet:
    def test_missing_header_returns_zero(self, fresh_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Page 13 Data"
        # No row with AUTO + MORTGAGE → header not found → 0 rows.
        ws.append(["some random text"])
        ws.append(["26:Q1", 1.0, 2.0])
        # Use the read-only workbook path the prod code does.
        buf = io.BytesIO()
        wb.save(buf)
        wb_ro = openpyxl.load_workbook(io.BytesIO(buf.getvalue()),
                                       read_only=True, data_only=True)
        n = hhdc._parse_flow_sheet(wb_ro["Page 13 Data"], 30, hhdc._now())
        assert n == 0

    def test_skips_rows_with_string_in_data_column(self, fresh_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Page 13 Data"
        ws.append(["Date", "AUTO", "CC", "MORTGAGE", "HELOC",
                   "STUDENT LOAN", "OTHER", "TOTAL"])
        # Row with a string in the AUTO column — that column must be skipped.
        ws.append(["26:Q1", "n/a", 8.0, 1.0, 0.5, 9.0, 2.0, 3.0])
        buf = io.BytesIO()
        wb.save(buf)
        wb_ro = openpyxl.load_workbook(io.BytesIO(buf.getvalue()),
                                       read_only=True, data_only=True)
        n = hhdc._parse_flow_sheet(wb_ro["Page 13 Data"], 30, hhdc._now())
        # 7 columns, but AUTO was a string → 6 stored values.
        assert n == 6
        with db.get_conn() as conn:
            ids = {r[0] for r in conn.execute(
                "SELECT series_id FROM metrics"
            )}
        assert "HHDC_FLOW30_AUTO" not in ids
        assert "HHDC_FLOW30_CC" in ids
